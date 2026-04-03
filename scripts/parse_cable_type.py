import openpyxl, json

wb = openpyxl.load_workbook(r"C:\Users\FREE\Desktop\cable type.xlsx", data_only=True)
print("Sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== SHEET: {sname} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 60:
            print(f"  ... ({ws.max_row - 60} more rows)")
            break
        vals = [str(v) if v is not None else '' for v in row]
        if any(v.strip() for v in vals):
            print(f"  Row{i+1}: {vals}")
